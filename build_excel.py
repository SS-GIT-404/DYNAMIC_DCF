"""
build_excel.py — Generate a *live* Excel DCF workbook for a ticker.

Every calculation in the workbook is a real Excel formula (e.g. =Rev0*(1+g_init)),
never a Python-computed constant, so the model genuinely recalculates when a reader
edits an assumption cell in Excel. Only two kinds of numbers are hard-coded:
  * reported historical actuals (SEC 10-K facts), and
  * the market-data snapshot (price / beta / shares at generation time).
Both are shown in blue and cited in an adjacent column.

Formatting follows standard financial-model conventions:
  * blue font  = hard-coded input (an actual, or a market snapshot)
  * yellow fill = a key assumption the reader is meant to edit (also blue font)
  * black font = a formula (calculated)
  * Arial throughout; currency / percent formats; negatives in (parentheses)

Input/output boundary
---------------------
The workbook is *generated* by running this script for a ticker — it is not a
live in-Excel query. See the module note printed by --help and the README for
why that is the right boundary (SEC XBRL is not a native Excel data source, and a
frozen snapshot is what makes the model auditable and reproducible).

Usage
-----
    python build_excel.py AAPL
    python build_excel.py AAPL --years 10 --out models/AAPL_DCF.xlsx
"""

from __future__ import annotations

import argparse
import os
from datetime import date
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, Reference

from sec_pull import Financials, pull_financials, DISPLAY_ORDER
from market_data import MarketData, fetch_market_data
from valuation import Assumptions, derive_assumptions

# --------------------------------------------------------------------------- #
# Styling constants
# --------------------------------------------------------------------------- #

FONT_NAME = "Arial"
BLUE = "FF0000CC"       # hard-coded inputs / actuals
BLACK = "FF000000"      # formulas
WHITE = "FFFFFFFF"
GREY = "FF666666"

YELLOW_FILL = PatternFill("solid", fgColor="FFFFF2CC")     # editable assumption
HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")     # section/column header
SUBHEAD_FILL = PatternFill("solid", fgColor="FFD9E1F2")    # sub-header band
RESULT_FILL = PatternFill("solid", fgColor="FFE2EFDA")     # highlighted result
WARN_FILL = PatternFill("solid", fgColor="FFFCE4E4")       # warning banner
DEFAULT_FILL = PatternFill("solid", fgColor="FFFFD7D7")    # fabricated (un-derived) input
RED = "FFC00000"                                           # warning text

# The [$-409] locale prefix pins US English number formatting, so thousands group
# as 182,447 rather than following the reader's locale (e.g. Indian 1,82,447).
# It belongs in the FIRST section only: repeating it in the negative section makes
# some readers (LibreOffice) print the code literally — "(-409343,338)". Verified
# empirically; the locale applies to the whole format string from one prefix.
FMT_USD = "[$-409]#,##0;(#,##0)"              # $ millions, negatives in parens
FMT_PRICE = "[$-409]$#,##0.00;($#,##0.00)"    # per-share
FMT_PCT = "[$-409]0.0%;(0.0%)"
FMT_PCT2 = "[$-409]0.00%;(0.00%)"
FMT_MULT = '[$-409]0.0"x"'
FMT_SH = "[$-409]#,##0"

THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _page_setup(ws, landscape: bool = True, fit_width: int = 1,
                fit_height: int = 0, title_rows: Optional[str] = None) -> None:
    """
    Make the sheet print sensibly: fit to one page wide so the citation column is
    never clipped, landscape, with header rows repeated on every page.
    """
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    if title_rows:
        ws.print_title_rows = title_rows


def _c(ws, ref, value=None, *, font_color=BLACK, bold=False, fill=None,
       fmt=None, align=None, italic=False, size=10, border=False, wrap=False):
    """Write a styled cell and return it."""
    cell = ws[ref]
    if value is not None:
        cell.value = value
    cell.font = Font(name=FONT_NAME, color=font_color, bold=bold, italic=italic, size=size)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = BORDER
    return cell


# --------------------------------------------------------------------------- #
# Assumptions sheet
# --------------------------------------------------------------------------- #

def _assumptions_sheet(wb: Workbook, a: Assumptions, fin: Financials,
                       market: MarketData, n: int,
                       warnings: Optional[List[str]] = None) -> Dict[str, str]:
    """Write the inputs/assumptions sheet; return {defined_name: 'Sheet!$B$r'}."""
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 52

    _c(ws, "A1", f"{fin.entity_name} ({fin.ticker}) — DCF Assumptions & Inputs",
       font_color=WHITE, bold=True, fill=HEADER_FILL, size=12)
    for col in "BCD":
        _c(ws, f"{col}1", "", fill=HEADER_FILL)
    _c(ws, "A2", f"Generated {date.today().isoformat()}  |  sector: {fin.sector}  |  "
                 f"horizon: {n} yrs  |  values in $ millions unless noted",
       font_color=GREY, italic=True, size=9)

    row = 3
    # ---- warning banner (model-fit + fabricated inputs) ------------------- #
    if warnings:
        _c(ws, f"A{row}", "READ THIS FIRST — model caveats", font_color=RED,
           bold=True, fill=WARN_FILL, size=10)
        for col in "BCD":
            _c(ws, f"{col}{row}", "", fill=WARN_FILL)
        row += 1
        for wmsg in warnings:
            ws.merge_cells(f"A{row}:D{row}")
            _c(ws, f"A{row}", f"!  {wmsg}", font_color=RED, fill=WARN_FILL,
               size=9, wrap=True)
            for col in "BCD":
                ws[f"{col}{row}"].fill = WARN_FILL
            ws.row_dimensions[row].height = 26
            row += 1
        row += 1

    _c(ws, f"A{row}", "Item", font_color=BLACK, bold=True, fill=SUBHEAD_FILL)
    _c(ws, f"B{row}", "Value", font_color=BLACK, bold=True, fill=SUBHEAD_FILL, align="center")
    _c(ws, f"C{row}", "Unit", font_color=BLACK, bold=True, fill=SUBHEAD_FILL, align="center")
    _c(ws, f"D{row}", "Source / basis", font_color=BLACK, bold=True, fill=SUBHEAD_FILL)
    row += 1

    names: Dict[str, str] = {}

    def section(title: str):
        nonlocal row
        _c(ws, f"A{row}", title, font_color=BLACK, bold=True, fill=SUBHEAD_FILL)
        for col in "BCD":
            _c(ws, f"{col}{row}", "", fill=SUBHEAD_FILL)
        row += 1

    def item(name: Optional[str], label: str, value, fmt, unit, source,
             kind: str = "assumption", attr: Optional[str] = None):
        """kind: 'assumption' (yellow+blue), 'input' (blue), 'calc' (black formula).

        `attr` names the Assumptions field so un-derivable ("defaulted") inputs can
        be filled red — a reader must never mistake a fabricated guess for a figure
        taken from the filings.
        """
        nonlocal row
        defaulted = bool(attr and a.is_defaulted(attr))
        _c(ws, f"A{row}", label, font_color=RED if defaulted else BLACK,
           bold=defaulted)
        if kind == "assumption":
            _c(ws, f"B{row}", value, font_color=BLUE,
               fill=DEFAULT_FILL if defaulted else YELLOW_FILL, fmt=fmt,
               align="center", border=True)
        elif kind == "input":
            _c(ws, f"B{row}", value, font_color=BLUE, fmt=fmt, align="center",
               border=True, fill=DEFAULT_FILL if defaulted else None)
        else:  # calc
            _c(ws, f"B{row}", value, font_color=BLACK, fmt=fmt, align="center", border=True)
        _c(ws, f"C{row}", unit, font_color=GREY, size=9, align="center")
        _c(ws, f"D{row}", source, font_color=RED if defaulted else GREY,
           size=9, wrap=True)
        if name:
            names[name] = f"Assumptions!$B${row}"
        row += 1

    src = a.sources
    section("FORECAST")
    item("Nyears", "Forecast horizon", n, "0", "yrs",
         "structural — regenerate the workbook to change the horizon", kind="input")
    item("g_init", "Revenue growth — year 1", a.initial_growth, FMT_PCT, "%/yr",
         src.get("initial_growth", "assumption"), attr="initial_growth")
    item("g_term", "Revenue growth — terminal", a.terminal_growth, FMT_PCT, "%/yr",
         src.get("terminal_growth", "assumption"), attr="terminal_growth")

    section("PROFITABILITY")
    item("m_start", "EBIT margin — latest actual", a.start_ebit_margin, FMT_PCT, "%",
         src.get("start_ebit_margin", "latest FY operating margin (10-K)"),
         attr="start_ebit_margin")
    item("m_target", "EBIT margin — steady state", a.target_ebit_margin, FMT_PCT, "%",
         src.get("target_ebit_margin", "assumption"), attr="target_ebit_margin")
    item("tax", "Effective tax rate", a.tax_rate, FMT_PCT, "%",
         src.get("tax_rate", "assumption"), attr="tax_rate")

    section("CAPITAL INTENSITY (% of revenue)")
    item("da_pct", "Depreciation & amortisation", a.da_pct, FMT_PCT, "%rev",
         src.get("da_pct", "assumption"), attr="da_pct")
    item("capex_pct", "Capital expenditure", a.capex_pct, FMT_PCT, "%rev",
         src.get("capex_pct", "assumption"), attr="capex_pct")
    item("nwc_pct", "Net working capital", a.nwc_pct, FMT_PCT, "%rev",
         src.get("nwc_pct", "assumption"), attr="nwc_pct")

    section("COST OF CAPITAL (CAPM / WACC)")
    item("rf", "Risk-free rate", a.risk_free_rate, FMT_PCT, "%",
         src.get("risk_free_rate", "assumption — 10y UST proxy"), attr="risk_free_rate")
    item("erp", "Equity risk premium", a.equity_risk_premium, FMT_PCT, "%",
         src.get("equity_risk_premium", "assumption"), attr="equity_risk_premium")
    item("beta", "Equity beta", a.beta, "0.00", "x",
         src.get("beta", "assumption"), attr="beta")
    item("kd_pre", "Pre-tax cost of debt", a.pretax_cost_of_debt, FMT_PCT, "%",
         src.get("pretax_cost_of_debt", "assumption"), attr="pretax_cost_of_debt")

    section("TERMINAL / EXIT")
    item("exit_mult", "Exit EV/EBITDA multiple", a.exit_ev_ebitda, FMT_MULT, "x",
         src.get("exit_ev_ebitda", "assumption"), attr="exit_ev_ebitda")

    section("CAPITAL STRUCTURE & MARKET (snapshot — hard-coded)")
    item("Rev0", "Base revenue (latest FY)", a.base_revenue / 1e6, FMT_USD, "$M",
         src.get("base_revenue", "SEC 10-K"), kind="input")
    item("NWC0", "Base net working capital", a.base_nwc / 1e6, FMT_USD, "$M",
         "AR + inventory - AP, latest FY (10-K)", kind="input")
    item("TotalDebt", "Total debt", a.total_debt / 1e6, FMT_USD, "$M",
         src.get("total_debt", "SEC 10-K"), kind="input")
    item("NetDebt", "Net debt", a.net_debt / 1e6, FMT_USD, "$M",
         "total debt - cash - short-term investments (10-K)", kind="input")
    item("Shares", "Shares outstanding", (a.shares or 0) / 1e6, FMT_SH, "M sh",
         f"{market.source if market.ok else 'SEC cover'} snapshot", kind="input")
    item("CurPrice", "Current share price", a.current_price, FMT_PRICE, "$",
         f"{market.source} snapshot {date.today().isoformat()}" if market.ok
         else "not available", kind="input")
    item("MktCap", "Market capitalisation", (a.market_cap or 0) / 1e6, FMT_USD, "$M",
         "price x shares (snapshot)", kind="input")

    # --- WACC build (formulas) -------------------------------------------- #
    section("WACC BUILD (calculated)")
    item("CostEquity", "Cost of equity", "=rf+beta*erp", FMT_PCT, "%",
         "CAPM: rf + beta x ERP", kind="calc")
    item("AfterTaxKd", "After-tax cost of debt", "=kd_pre*(1-tax)", FMT_PCT, "%",
         "kd x (1 - tax)", kind="calc")
    item("WgtEquity", "Weight of equity", "=MktCap/(MktCap+TotalDebt)", FMT_PCT, "%",
         "E / (E + D)", kind="calc")
    item("WgtDebt", "Weight of debt", "=TotalDebt/(MktCap+TotalDebt)", FMT_PCT, "%",
         "D / (E + D)", kind="calc")
    item("WACC", "WACC", "=WgtEquity*CostEquity+WgtDebt*AfterTaxKd", FMT_PCT, "%",
         "weighted average cost of capital", kind="calc")

    # --- colour-convention legend ------------------------------------------ #
    row += 1
    _c(ws, f"A{row}", "LEGEND — model conventions", bold=True, fill=SUBHEAD_FILL)
    for col in "BCD":
        _c(ws, f"{col}{row}", "", fill=SUBHEAD_FILL)
    row += 1
    legend = [
        (YELLOW_FILL, BLUE, "Yellow fill, blue font",
         "Key assumption — intended for you to edit. Drives the whole model."),
        (None, BLUE, "Blue font (no fill)",
         "Hard-coded input: a reported SEC actual or a market snapshot."),
        (None, BLACK, "Black font",
         "Formula — calculated live by Excel. Do not overwrite."),
        (DEFAULT_FILL, RED, "Red fill / red text",
         "Could NOT be derived from this company's filings — a generic default. "
         "Verify or replace before relying on the output."),
    ]
    for fill, color, label_text, desc in legend:
        _c(ws, f"A{row}", label_text, font_color=color, fill=fill, size=9, border=True)
        ws.merge_cells(f"B{row}:D{row}")
        _c(ws, f"B{row}", desc, font_color=GREY, size=9, wrap=True)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    _page_setup(ws)
    return names


# --------------------------------------------------------------------------- #
# Historical sheet (SEC actuals — hard-coded, cited)
# --------------------------------------------------------------------------- #

def _historical_sheet(wb: Workbook, fin: Financials) -> None:
    ws = wb.create_sheet("Historical")
    ws.column_dimensions["A"].width = 24
    years = fin.fiscal_years
    _c(ws, "A1", f"{fin.ticker} — Historical financials (SEC EDGAR 10-K, $M)",
       font_color=WHITE, bold=True, fill=HEADER_FILL, size=12)
    for j in range(len(years) + 1):
        _c(ws, f"{get_column_letter(j + 2)}1", "", fill=HEADER_FILL)
    _c(ws, "A2", "Reported actuals — hard-coded (blue) and cited; not part of the "
                 "live formula chain.", font_color=GREY, italic=True, size=9)

    _c(ws, "A3", "Line item ($M)", bold=True, fill=SUBHEAD_FILL)
    for j, y in enumerate(years):
        col = get_column_letter(j + 2)
        _c(ws, f"{col}3", f"FY{y}", bold=True, fill=SUBHEAD_FILL, align="center")
        ws.column_dimensions[col].width = 11
    src_col = get_column_letter(len(years) + 2)
    _c(ws, f"{src_col}3", "Source", bold=True, fill=SUBHEAD_FILL)
    ws.column_dimensions[src_col].width = 40

    row = 4
    for name in DISPLAY_ORDER:
        status = fin.item_status.get(name, "not_found")
        _c(ws, f"A{row}", name)
        for j, y in enumerate(years):
            col = get_column_letter(j + 2)
            v = fin.line_items.get(name, {}).get(y)
            if v is None:
                label = "n/a" if status == "not_applicable" else "--"
                _c(ws, f"{col}{row}", label, font_color=GREY, align="center", size=9)
            else:
                _c(ws, f"{col}{row}", v / 1e6, font_color=BLUE, fmt=FMT_USD, align="right")
        tag = fin.tags_used.get(name, "" if status == "ok" else status.replace("_", " "))
        _c(ws, f"{src_col}{row}", f"XBRL: {tag}" if tag else "", font_color=GREY, size=8, wrap=True)
        row += 1

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    _page_setup(ws, title_rows="3:3")


# --------------------------------------------------------------------------- #
# DCF sheet (all live formulas)
# --------------------------------------------------------------------------- #

def _dcf_sheet(wb: Workbook, n: int) -> Dict[str, str]:
    """Write the forecast + valuation bridge as live formulas. Returns key cell refs."""
    ws = wb.create_sheet("DCF")
    ws.column_dimensions["A"].width = 31
    first = 3                       # first forecast column index (C)
    cols = [get_column_letter(first + t) for t in range(n)]   # C..(C+n-1)
    base_col = get_column_letter(first - 1)                   # B (base / FY0)

    _c(ws, "A1", "Discounted Cash Flow — unlevered FCFF ($M)",
       font_color=WHITE, bold=True, fill=HEADER_FILL, size=12)
    for t in range(n + 1):
        _c(ws, f"{get_column_letter(first - 1 + t)}1", "", fill=HEADER_FILL)
    _c(ws, "A2", "Black = formula (recalculates). Edit assumptions on the "
                 "Assumptions tab.", font_color=GREY, italic=True, size=9)

    # header rows
    _c(ws, f"A3", "($ millions)", bold=True, fill=SUBHEAD_FILL)
    _c(ws, f"{base_col}3", "Base FY0", bold=True, fill=SUBHEAD_FILL, align="center")
    for t, col in enumerate(cols, start=1):
        _c(ws, f"{col}3", f"Year {t}", bold=True, fill=SUBHEAD_FILL, align="center")
        ws.column_dimensions[col].width = 12
    ws.column_dimensions[base_col].width = 12

    R = {}   # label -> row number
    r = 4

    def label(row, text, **kw):
        _c(ws, f"A{row}", text, **kw)

    # t (period index)
    R["t"] = r
    label(r, "Period (t)", font_color=GREY, size=9)
    _c(ws, f"{base_col}{r}", 0, font_color=GREY, align="center", size=9)
    for t, col in enumerate(cols, start=1):
        _c(ws, f"{col}{r}", t, font_color=GREY, align="center", size=9)
    r += 1

    # revenue growth (faded)
    R["g"] = r
    label(r, "Revenue growth")
    for t, col in enumerate(cols, start=1):
        f = (f"=g_init+(g_term-g_init)*({col}{R['t']}-1)/(Nyears-1)" if n > 1
             else "=g_init")
        _c(ws, f"{col}{r}", f, fmt=FMT_PCT, align="center")
    r += 1

    # revenue
    R["rev"] = r
    label(r, "Revenue", bold=True)
    _c(ws, f"{base_col}{r}", "=Rev0", fmt=FMT_USD, align="right")
    prev = base_col
    for t, col in enumerate(cols, start=1):
        _c(ws, f"{col}{r}", f"={prev}{r}*(1+{col}{R['g']})", fmt=FMT_USD, align="right")
        prev = col
    r += 1

    # ebit margin (ramped)
    R["margin"] = r
    label(r, "EBIT margin")
    _c(ws, f"{base_col}{r}", "=m_start", fmt=FMT_PCT, align="center")
    for t, col in enumerate(cols, start=1):
        _c(ws, f"{col}{r}", f"=m_start+(m_target-m_start)*{col}{R['t']}/Nyears",
           fmt=FMT_PCT, align="center")
    r += 1

    def driver(key, text, formula_of_col, bold=False):
        nonlocal r
        R[key] = r
        label(r, text, bold=bold)
        for col in cols:
            _c(ws, f"{col}{r}", formula_of_col(col), fmt=FMT_USD, align="right")
        r += 1

    driver("ebit", "EBIT", lambda c: f"={c}{R['rev']}*{c}{R['margin']}", bold=True)
    driver("nopat", "NOPAT = EBIT x (1-tax)", lambda c: f"={c}{R['ebit']}*(1-tax)")
    driver("da", "(+) D&A", lambda c: f"={c}{R['rev']}*da_pct")
    driver("capex", "(-) Capex", lambda c: f"={c}{R['rev']}*capex_pct")

    # NWC + change
    R["nwc"] = r
    label(r, "Net working capital")
    _c(ws, f"{base_col}{r}", "=NWC0", fmt=FMT_USD, align="right")
    for col in cols:
        _c(ws, f"{col}{r}", f"={col}{R['rev']}*nwc_pct", fmt=FMT_USD, align="right")
    r += 1

    R["dnwc"] = r
    label(r, "(-) Change in NWC")
    prev = base_col
    for col in cols:
        _c(ws, f"{col}{r}", f"={col}{R['nwc']}-{prev}{R['nwc']}", fmt=FMT_USD, align="right")
        prev = col
    r += 1

    # FCFF
    R["fcff"] = r
    label(r, "Unlevered FCFF", bold=True)
    for col in cols:
        _c(ws, f"{col}{r}",
           f"={col}{R['nopat']}+{col}{R['da']}-{col}{R['capex']}-{col}{R['dnwc']}",
           fmt=FMT_USD, align="right", bold=True)
    r += 1

    # discount factor + PV
    R["df"] = r
    label(r, "Discount factor")
    for col in cols:
        _c(ws, f"{col}{r}", f"=1/(1+WACC)^{col}{R['t']}", fmt="0.000", align="center")
    r += 1

    R["pv"] = r
    label(r, "PV of FCFF", bold=True)
    for col in cols:
        _c(ws, f"{col}{r}", f"={col}{R['fcff']}*{col}{R['df']}", fmt=FMT_USD, align="right", bold=True)
    r += 1

    last_col = cols[-1]
    fcff_range = f"{cols[0]}{R['fcff']}:{cols[-1]}{R['fcff']}"
    pv_range = f"{cols[0]}{R['pv']}:{cols[-1]}{R['pv']}"

    # ---- valuation bridge ------------------------------------------------- #
    r += 1
    _c(ws, f"A{r}", "VALUATION BRIDGE", bold=True, fill=SUBHEAD_FILL)
    _c(ws, f"B{r}", "Gordon", bold=True, fill=SUBHEAD_FILL, align="center")
    _c(ws, f"C{r}", "Exit mult.", bold=True, fill=SUBHEAD_FILL, align="center")
    r += 1

    def bridge(label_text, formula_b, formula_c=None, fmt=FMT_USD, name=None,
               bold=False, fill=None):
        nonlocal r
        _c(ws, f"A{r}", label_text, bold=bold)
        _c(ws, f"B{r}", formula_b, fmt=fmt, align="right", bold=bold, fill=fill)
        if formula_c is not None:
            _c(ws, f"C{r}", formula_c, fmt=fmt, align="right", bold=bold, fill=fill)
        ref = f"DCF!$B${r}"
        if name:
            R[name] = r
        r += 1
        return ref

    refs = {}
    refs["sumpv"] = bridge("PV of explicit FCFF", f"=SUM({pv_range})", f"=SUM({pv_range})", name="sumpv")
    # terminal values
    refs["tv"] = bridge("Terminal value (undiscounted)",
                        f"={last_col}{R['fcff']}*(1+g_term)/(WACC-g_term)",
                        f"=({last_col}{R['ebit']}+{last_col}{R['da']})*exit_mult",
                        name="tv")
    tv_row = R["tv"]
    sumpv_row = R["sumpv"]
    refs["pvtv"] = bridge("PV of terminal value",
                          f"=B{tv_row}*{last_col}{R['df']}",
                          f"=C{tv_row}*{last_col}{R['df']}", name="pvtv")
    pvtv_row = R["pvtv"]
    refs["ev"] = bridge("Enterprise value",
                        f"=B{sumpv_row}+B{pvtv_row}",
                        f"=C{sumpv_row}+C{pvtv_row}", name="ev", bold=True)
    ev_row = R["ev"]
    refs["nd"] = bridge("(-) Net debt", "=NetDebt", "=NetDebt", name="nd")
    nd_row = R["nd"]
    refs["eq"] = bridge("Equity value",
                        f"=B{ev_row}-B{nd_row}", f"=C{ev_row}-C{nd_row}",
                        name="eq", bold=True)
    eq_row = R["eq"]
    refs["price"] = bridge("Implied share price",
                           f"=B{eq_row}/Shares", f"=C{eq_row}/Shares",
                           fmt=FMT_PRICE, name="price", bold=True, fill=RESULT_FILL)
    price_row = R["price"]
    bridge("Current share price", "=CurPrice", "=CurPrice", fmt=FMT_PRICE)
    refs["upside"] = bridge("Upside / (downside)",
                            f"=B{price_row}/CurPrice-1", f"=C{price_row}/CurPrice-1",
                            fmt=FMT_PCT, name="upside", bold=True)

    # cross-checks
    r += 1
    _c(ws, f"A{r}", "CROSS-CHECKS", bold=True, fill=SUBHEAD_FILL)
    _c(ws, f"B{r}", "", fill=SUBHEAD_FILL)
    r += 1
    _c(ws, f"A{r}", "Terminal value % of EV (Gordon)")
    _c(ws, f"B{r}", f"=B{pvtv_row}/B{ev_row}", fmt=FMT_PCT, align="right")
    r += 1
    _c(ws, f"A{r}", "Gordon TV implied EV/EBITDA")
    _c(ws, f"B{r}", f"=B{tv_row}/({last_col}{R['ebit']}+{last_col}{R['da']})",
       fmt=FMT_MULT, align="right")
    r += 1

    # forecast FCFF bar chart
    chart = BarChart()
    chart.title = "Unlevered FCFF forecast ($M)"
    chart.type = "col"
    chart.legend = None
    chart.height = 6.5
    chart.width = 13
    data = Reference(ws, min_col=first, max_col=first + n - 1, min_row=R["fcff"], max_row=R["fcff"])
    cats = Reference(ws, min_col=first, max_col=first + n - 1, min_row=3, max_row=3)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{r + 1}")

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    _page_setup(ws)
    R["_cols"] = cols
    R["_base_col"] = base_col
    R["price_row"] = price_row
    return R


# --------------------------------------------------------------------------- #
# Sensitivity sheet (live formulas: WACC x terminal growth)
# --------------------------------------------------------------------------- #

def _sensitivity_sheet(wb: Workbook, n: int, dcf: Dict[str, str]) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.column_dimensions["A"].width = 12
    cols = dcf["_cols"]
    fcff_row = dcf["fcff"]
    df_last = None
    wacc_deltas = [-0.015, -0.010, -0.005, 0.0, 0.005, 0.010, 0.015]
    g_deltas = [-0.010, -0.005, 0.0, 0.005, 0.010]

    _c(ws, "A1", "Sensitivity — implied share price (Gordon): WACC x terminal growth",
       font_color=WHITE, bold=True, fill=HEADER_FILL, size=12)
    # Start at column B: filling from A would overwrite the title just written.
    for j in range(len(g_deltas) + 1):
        _c(ws, f"{get_column_letter(j + 2)}1", "", fill=HEADER_FILL)
    _c(ws, "A2", "Every cell is a live formula off the DCF FCFF stream — recalculates "
                 "with the assumptions.", font_color=GREY, italic=True, size=9)

    corner_r = 4
    _c(ws, f"A{corner_r}", "=DCF!$B$%d" % dcf["price_row"], font_color=BLACK,
       bold=True, fmt=FMT_PRICE, align="center", fill=RESULT_FILL, border=True)
    _c(ws, f"A{corner_r - 1}", "WACC \\ g", font_color=GREY, size=9, align="center")

    # column headers: terminal growth = g_term + delta
    g_cols = [get_column_letter(2 + j) for j in range(len(g_deltas))]
    for j, (gc, d) in enumerate(zip(g_cols, g_deltas)):
        _c(ws, f"{gc}{corner_r}", f"=g_term+({d})", fmt=FMT_PCT, bold=True,
           align="center", fill=SUBHEAD_FILL, border=True)
        ws.column_dimensions[gc].width = 11

    # row headers: WACC = WACC + delta
    for i, d in enumerate(wacc_deltas):
        rr = corner_r + 1 + i
        _c(ws, f"A{rr}", f"=WACC+({d})", fmt=FMT_PCT2, bold=True, align="center",
           fill=SUBHEAD_FILL, border=True)

    # interior cells: implied price given (wacc=$A{rr}, g={gc}${corner_r})
    for i in range(len(wacc_deltas)):
        rr = corner_r + 1 + i
        w_ref = f"$A{rr}"
        for gc in g_cols:
            g_ref = f"{gc}${corner_r}"
            pv_terms = "+".join(
                f"DCF!${cols[t]}${fcff_row}/(1+{w_ref})^{t + 1}" for t in range(n))
            tv_term = (f"DCF!${cols[-1]}${fcff_row}*(1+{g_ref})/"
                       f"({w_ref}-{g_ref})/(1+{w_ref})^{n}")
            formula = f"=(({pv_terms})+({tv_term})-NetDebt)/Shares"
            _c(ws, f"{gc}{rr}", formula, fmt=FMT_PRICE, align="center", border=True)

    _c(ws, f"A{corner_r + len(wacc_deltas) + 2}",
       "Center cell (boxed value on DCF) = base case. Blue headers recalc off WACC and "
       "terminal g on the Assumptions tab.", font_color=GREY, italic=True, size=9)
    ws.sheet_view.showGridLines = False
    _page_setup(ws)


# --------------------------------------------------------------------------- #
# Assemble workbook
# --------------------------------------------------------------------------- #

def build_workbook(ticker: str, forecast_years: int = 5, out_path: Optional[str] = None,
                   fin: Optional[Financials] = None, market: Optional[MarketData] = None,
                   assumptions: Optional[Assumptions] = None) -> str:
    fin = fin or pull_financials(ticker)
    market = market if market is not None else fetch_market_data(ticker)
    a = assumptions or derive_assumptions(fin, market, forecast_years=forecast_years)
    n = a.forecast_years

    # Run the engine purely to collect its caveats, so the workbook carries the
    # same warnings the Python/Streamlit surfaces show.
    from valuation import run_dcf
    caveats = run_dcf(a, ticker=fin.ticker, entity_name=fin.entity_name,
                      sector=fin.sector).warnings

    wb = Workbook()
    wb.remove(wb.active)   # drop default sheet

    names = _assumptions_sheet(wb, a, fin, market, n, warnings=caveats)
    _historical_sheet(wb, fin)
    dcf = _dcf_sheet(wb, n)
    _sensitivity_sheet(wb, n, dcf)

    # register defined names (workbook scope)
    for nm, ref in names.items():
        wb.defined_names.add(DefinedName(nm, attr_text=ref))

    # order sheets: Assumptions, DCF, Sensitivity, Historical
    wb.move_sheet("Historical", offset=2)

    out_path = out_path or os.path.join("models", f"{fin.ticker}_DCF_Model.xlsx")
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("ticker")
    p.add_argument("--years", type=int, default=5)
    p.add_argument("--out", default=None)
    p.add_argument("--no-market", action="store_true")
    args = p.parse_args(argv)

    fin = pull_financials(args.ticker)
    market = MarketData(ticker=args.ticker.upper()) if args.no_market \
        else fetch_market_data(args.ticker)
    path = build_workbook(args.ticker, forecast_years=args.years, out_path=args.out,
                          fin=fin, market=market)
    print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
