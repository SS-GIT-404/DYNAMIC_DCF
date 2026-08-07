"""
verify_model.py — Cross-check the Excel workbook against the Python engine.

The Excel workbook and the Python engine are two independent implementations of the
same DCF. This script builds (or reads) a workbook, recalculates it with a real
spreadsheet engine, and asserts that the headline outputs agree with `valuation.py`
to within a tight tolerance.

That agreement is the actual guarantee behind the workbook: not just "no #REF!
errors", but "the live Excel formulas reproduce the model".

Usage
-----
    python verify_model.py AAPL
    python verify_model.py AAPL JPM O --years 5
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import List, Optional

from sec_pull import pull_financials
from market_data import fetch_market_data
from valuation import derive_assumptions, run_dcf
from build_excel import build_workbook
from recalc_check import recalc_libreoffice, recalc_formulas, scan_errors, lookup

def find_cell(path: str, sheet: str, label: str, col: str = "B") -> Optional[str]:
    """
    Locate a value cell by its row label, e.g. ('DCF', 'Implied share price') -> 'DCF!B25'.

    Row positions shift as the layout changes (a warning banner, an extra section),
    so the verifier addresses cells by label rather than by hard-coded coordinates.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    target = label.strip().lower()
    exact, prefix = None, None
    for row in ws.iter_rows(min_col=1, max_col=1):
        val = row[0].value
        if not val:
            continue
        text = str(val).strip().lower()
        # An exact label wins over a prefix hit, so looking up "WACC" finds the
        # WACC row itself rather than the "WACC BUILD (calculated)" section header.
        if text == target and exact is None:
            exact = f"{sheet}!{col}{row[0].row}"
        elif text.startswith(target) and prefix is None:
            prefix = f"{sheet}!{col}{row[0].row}"
    return exact or prefix


def verify(ticker: str, years: int = 5, engine: str = "auto",
           rebuild: bool = True, tol_price: float = 0.05,
           tol_rate: float = 1e-4) -> bool:
    print("=" * 74)
    print(f"VERIFY {ticker}  (horizon {years}y)")
    print("=" * 74)

    fin = pull_financials(ticker)
    market = fetch_market_data(ticker)
    a = derive_assumptions(fin, market, forecast_years=years)
    res = run_dcf(a, ticker=fin.ticker, entity_name=fin.entity_name, sector=fin.sector)

    path = os.path.join("models", f"{fin.ticker}_DCF_Model.xlsx")
    if rebuild or not os.path.exists(path):
        path = build_workbook(ticker, forecast_years=years, out_path=path,
                              fin=fin, market=market, assumptions=a)
    print(f"  workbook      : {path}")

    values, engine_name = None, ""
    for fn in ([recalc_libreoffice, recalc_formulas] if engine == "auto"
               else [recalc_libreoffice] if engine == "libreoffice"
               else [recalc_formulas]):
        try:
            values, engine_name = fn(path)
            break
        except Exception as exc:  # noqa: BLE001
            print(f"  ({fn.__name__} unavailable: {exc})")
    if values is None:
        print("  ERROR: no recalculation engine available")
        return False
    print(f"  engine        : {engine_name}")

    errors = scan_errors(values)
    print(f"  formula errors: {len(errors)}")
    for ref, val in errors[:10]:
        print(f"      {ref} -> {val}")

    # Excel carries $ millions; the Python engine carries raw dollars.
    def at(sheet, label, col="B"):
        ref = find_cell(path, sheet, label, col)
        return lookup(values, ref) if ref else None

    checks = [
        ("WACC", at("Assumptions", "WACC"), res.wacc.wacc, tol_rate, "rate"),
        ("Cost of equity", at("Assumptions", "Cost of equity"),
         res.wacc.cost_of_equity, tol_rate, "rate"),
        ("EV (Gordon)", at("DCF", "Enterprise value"), res.ev_gordon / 1e6, 1.0, "usd_m"),
        ("EV (exit mult)", at("DCF", "Enterprise value", "C"), res.ev_exit / 1e6, 1.0, "usd_m"),
        ("Price (Gordon)", at("DCF", "Implied share price"), res.price_gordon, tol_price, "price"),
        ("Price (exit mult)", at("DCF", "Implied share price", "C"),
         res.price_exit, tol_price, "price"),
    ]
    if res.upside_gordon is not None:
        checks.append(("Upside (Gordon)", at("DCF", "Upside"),
                       res.upside_gordon, 1e-3, "rate"))

    print(f"  {'output':<20}{'Excel':>16}{'Python':>16}{'diff':>12}   status")
    ok_all = not errors
    for label, xl_val, py_val, tol, kind in checks:
        if xl_val is None or not isinstance(xl_val, (int, float)):
            print(f"  {label:<20}{'n/a':>16}{py_val:>16,.4f}{'':>12}   MISSING")
            ok_all = False
            continue
        diff = abs(float(xl_val) - float(py_val))
        ok = diff <= tol
        ok_all = ok_all and ok
        if kind == "rate":
            xs, ps, ds = f"{xl_val*100:,.3f}%", f"{py_val*100:,.3f}%", f"{diff*100:,.4f}pp"
        elif kind == "price":
            xs, ps, ds = f"${xl_val:,.2f}", f"${py_val:,.2f}", f"${diff:,.3f}"
        else:
            xs, ps, ds = f"{xl_val:,.0f}", f"{py_val:,.0f}", f"{diff:,.1f}"
        print(f"  {label:<20}{xs:>16}{ps:>16}{ds:>12}   {'OK' if ok else 'FAIL'}")

    print(f"\n  {ticker}: {'PASS' if ok_all else 'FAIL'}\n")
    return ok_all


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Cross-check Excel workbook vs Python engine.")
    p.add_argument("tickers", nargs="+")
    p.add_argument("--years", type=int, default=5)
    p.add_argument("--engine", choices=["auto", "libreoffice", "formulas"], default="auto")
    p.add_argument("--no-rebuild", action="store_true")
    args = p.parse_args(argv)

    results = {}
    for tk in args.tickers:
        try:
            results[tk] = verify(tk, years=args.years, engine=args.engine,
                                 rebuild=not args.no_rebuild)
        except Exception as exc:  # noqa: BLE001
            print(f"  {tk}: ERROR — {exc}\n")
            results[tk] = False

    print("=" * 74)
    for tk, ok in results.items():
        print(f"  {tk:<8} {'PASS' if ok else 'FAIL'}")
    print("=" * 74)
    return 0 if all(results.values()) else 1


if __name__ == "__main__":
    raise SystemExit(main())
